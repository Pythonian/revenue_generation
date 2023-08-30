from datetime import datetime

import openpyxl
from django.contrib import messages
from django.db.models import Count, Sum
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.db.models import Q
from django.utils import timezone
from django.views.decorators.http import require_POST
from openpyxl.utils import get_column_letter

from .forms import (RevenueCategoryForm, RevenueSourceForm,
                    RevenueTransactionForm, TaxPayerForm)
from .models import (RevenueCategory, RevenueSource, RevenueTransaction,
                     TaxPayer)
from .utils import mk_paginator


def index(request):
    return render(request, 'index.html')

def home(request):
    current_year = datetime.today().year
    current_month = datetime.today().month
    prev_month = current_month - 1 if current_month > 1 else 12
    last_year = current_year - 1
    # Initialize with zeros for all months
    revenue_data = [0] * 12

    transactions = RevenueTransaction.objects.filter(
        created__year=current_year)
    latest_transactions = RevenueTransaction.objects.filter(
        created__year=current_year)[:5]
    for transaction in transactions:
        revenue_data[transaction.created.month - 1] += transaction.amount

    # Calculate total revenue for the current year
    current_year = datetime.today().year
    total_revenue_this_year = (
        RevenueTransaction.objects.filter(
            created__year=current_year).aggregate(
                total=Sum("amount")
        )["total"]
        or 0
    )

    # Count total number of revenue transactions in the current year
    total_revenue_transactions_this_year = RevenueTransaction.objects.filter(
        created__year=current_year
    ).count()

    # Count total number of revenue transactions in the current year
    total_revenue_transactions_last_year = RevenueTransaction.objects.filter(
        created__year=last_year
    ).count()

    # Count total number of registered Tax Payers in the database
    total_taxpayers = TaxPayer.objects.all().count()
    total_taxpayers_registered_this_year = TaxPayer.objects.filter(
        created__year=current_year).count()

    # Calculate average revenue per transaction in the current year
    average_revenue_per_transaction = (
        total_revenue_this_year / total_revenue_transactions_this_year if total_revenue_transactions_this_year > 0 else 0
    )

    total_revenue_prev_month = (
        RevenueTransaction.objects.filter(
            created__year=current_year, created__month=prev_month
        ).aggregate(Sum("amount"))["amount__sum"]
        or 0
    )

    top_revenue_sources = RevenueSource.objects.annotate(
        transaction_count=Count("revenuetransaction")
    ).order_by("-transaction_count")[:5]

    max_transaction_count = top_revenue_sources.first().transaction_count

    top_revenue_sources_with_progress = []
    for source in top_revenue_sources:
        progress = (source.transaction_count / max_transaction_count) * 100
        top_revenue_sources_with_progress.append(
            {
                "source": source,
                "progress": progress,
            }
        )

    # Calculate revenue from last year
    last_year = datetime.today().year - 1
    revenue_last_year = RevenueTransaction.objects.filter(
            created__year=last_year
        ).aggregate(
            total_revenue=Sum("amount")
        )["total_revenue"] or 0

    # Calculate revenue for this year
    current_year = datetime.today().year
    revenue_this_year = RevenueTransaction.objects.filter(
            created__year=current_year
        ).aggregate(
            total_revenue=Sum("amount")
        )["total_revenue"] or 0

    # Calculate the percentage increase or decrease in revenue
    if revenue_last_year == 0:
        revenue_percentage_change = 100  # Handle the case when last year's revenue is zero
    else:
        revenue_percentage_change = (
            (revenue_this_year - revenue_last_year) / revenue_last_year
        ) * 100

    # Determine whether it's an increase or decrease in revenue
    revenue_change_type = "increase" if revenue_percentage_change > 0 else "decrease"
    revenue_percentage_change = abs(revenue_percentage_change)

    # Calculate the percentage increase or descrease in transactions
    if total_revenue_transactions_last_year == 0:
        transaction_percentage_change = 100
    else:
        transaction_percentage_change = (
            (total_revenue_transactions_this_year - total_revenue_transactions_last_year) / 
            total_revenue_transactions_last_year
        ) * 100

    # Determine whether it's an increase or decrease in transaction
    transaction_change_type = "increase" if transaction_percentage_change > 0 else "decrease"
    transaction_percentage_change = abs(transaction_percentage_change)

    # Remove months with zero revenue
    non_zero_revenue_data = [
        (month, revenue) for month, revenue in enumerate(revenue_data, start=1) if revenue > 0
    ]

    # Calculate total number of transactions for each revenue source
    revenue_sources_with_transaction_count = RevenueSource.objects.annotate(
        transaction_count=Count("revenuetransaction")
    )

    context = {
        "total_revenue_this_year": total_revenue_this_year,
        "total_revenue_transactions_this_year": total_revenue_transactions_this_year,
        "total_taxpayers": total_taxpayers,
        "latest_transactions": latest_transactions,
        "average_revenue_per_transaction": average_revenue_per_transaction,
        "revenue_data": revenue_data,
        "total_revenue_prev_month": total_revenue_prev_month,
        "top_revenue_sources": top_revenue_sources_with_progress,
        "revenue_last_year": revenue_last_year,
        "revenue_this_year": revenue_this_year,
        "revenue_percentage_change": revenue_percentage_change,
        "revenue_change_type": revenue_change_type,
        "non_zero_revenue_data": non_zero_revenue_data,
        "total_revenue_transactions_last_year": total_revenue_transactions_last_year,
        "transaction_change_type": transaction_change_type,
        "total_revenue_transactions_last_year": total_revenue_transactions_last_year,
        "transaction_percentage_change": transaction_percentage_change,
        "total_taxpayers_registered_this_year": total_taxpayers_registered_this_year,
        "revenue_sources_with_transaction_count": revenue_sources_with_transaction_count,
    }

    return render(request, "home.html", context)


###########################################################
#
# VIEW FOR TAX PAYERS
#
###########################################################

def tax_payer_list(request):
    tax_payers = TaxPayer.objects.all()
    search_query = request.GET.get('q')

    if search_query:
        tax_payers = tax_payers.filter(
            Q(name__icontains=search_query) |   # Search by name
            Q(tin__icontains=search_query) |    # Search by TIN
            Q(email__icontains=search_query) |  # Search by email
            Q(phone__icontains=search_query)    # Search by phone
        )

    tax_payers = mk_paginator(request, tax_payers, 12)

    if request.method == "POST":
        form = TaxPayerForm(request.POST)
        if form.is_valid():
            tax_payer = form.save(commit=False)
            tax_payer.created = timezone.now()
            tax_payer.save()
            messages.success(
                request, f"{tax_payer.name} has been successfully created.")
            return redirect("tax_payer_detail", tax_payer.id)
    else:
        form = TaxPayerForm()

    # Excel file generation and download
    if request.GET.get("download_excel"):
        response = HttpResponse(content_type="application/ms-excel")
        response["Content-Disposition"] = "attachment; filename=tax_payers.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Tax Payers"

        # Add header row
        header = ["Tax Payer", "Tax Identification Number", "Phone", "Email"]
        for col_num, header_text in enumerate(header, 1):
            col_letter = get_column_letter(col_num)
            ws[f"{col_letter}1"] = header_text

        # Add tax_payer data
        for row_num, tax_payer in enumerate(tax_payers, 2):
            ws[f"A{row_num}"] = tax_payer.name
            ws[f"B{row_num}"] = tax_payer.tin
            ws[f"C{row_num}"] = tax_payer.phone
            ws[f"D{row_num}"] = tax_payer.email

        wb.save(response)
        return response

    return render(request,
                  "tax_payer_list.html",
                  {"tax_payers": tax_payers,
                   "form": form})


def tax_payer_detail(request, id):
    tax_payer = get_object_or_404(TaxPayer, id=id)
    transactions = RevenueTransaction.objects.filter(tax_payer=tax_payer)
    transactions = mk_paginator(request, transactions, 30)

    if request.method == "POST":
        form = TaxPayerForm(request.POST, instance=tax_payer)
        if form.is_valid():
            form.save()
            messages.success(
                request, f"{tax_payer.name} has been successfully updated.")
            return redirect("tax_payer_detail", tax_payer.id)
    else:
        form = TaxPayerForm(instance=tax_payer)

    # Excel file generation and download
    if request.GET.get("download_excel"):
        response = HttpResponse(content_type="application/ms-excel")
        response["Content-Disposition"] = f"attachment; filename={tax_payer.name}_transactions.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Transactions"

        # Add header row
        header = ["Transaction ID", "Revenue Source", "Amount", "Note", "Created"]
        for col_num, header_text in enumerate(header, 1):
            col_letter = get_column_letter(col_num)
            ws[f"{col_letter}1"] = header_text

        # Add transaction data
        for row_num, transaction in enumerate(transactions, 2):
            ws[f"A{row_num}"] = transaction.id
            ws[f"B{row_num}"] = str(transaction.revenue_source)
            ws[f"C{row_num}"] = transaction.amount
            ws[f"D{row_num}"] = transaction.note
            created_naive = timezone.localtime(transaction.created).replace(tzinfo=None)
            ws[f"E{row_num}"] = created_naive

        wb.save(response)
        return response

    return render(request,
                  "tax_payer_detail.html",
                  {"tax_payer": tax_payer,
                   "transactions": transactions,
                   "form": form})


@require_POST
def tax_payer_delete(request):
    if request.method == "POST":
        tax_payer_id = request.POST.get("tax_payer_id")
        tax_payer = get_object_or_404(TaxPayer, id=tax_payer_id)
        tax_payer.delete()
        messages.success(request, "Tax Payer successfully deleted.")
        return redirect(reverse("tax_payer_list"))


###########################################################
#
# VIEW FOR REVENUE TRANSACTIONS
#
###########################################################

def transaction_list(request):
    transactions = RevenueTransaction.objects.all()
    search_query = request.GET.get('q')

    if search_query:
        transactions = transactions.filter(
            Q(tax_payer__name__icontains=search_query) |
            Q(tax_payer__tin__icontains=search_query) |
            Q(note__icontains=search_query) |
            Q(revenue_source__name__icontains=search_query)
        )
    transactions = mk_paginator(request, transactions, 50)

    if request.method == "POST":
        form = RevenueTransactionForm(request.POST)
        if form.is_valid():
            transaction = form.save(commit=False)
            transaction.created = timezone.now()
            transaction.save()
            messages.success(
                request, "Revenue Transaction has been successfully created.")
            return redirect("transaction_list")
    else:
        form = RevenueTransactionForm()

    # Excel file generation and download
    if request.GET.get("download_excel"):
        response = HttpResponse(content_type="application/ms-excel")
        response["Content-Disposition"] = "attachment; filename=RevenueTransactions.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Revenue Transactions"

        # Add header row
        header = [
            "Transaction ID", "Tax Payer",
            "Tax Identification Number", "Revenue Type",
            "Amount", "Note", "Created"]
        for col_num, header_text in enumerate(header, 1):
            col_letter = get_column_letter(col_num)
            ws[f"{col_letter}1"] = header_text

        # Add transaction data
        for row_num, transaction in enumerate(transactions, 2):
            ws[f"A{row_num}"] = transaction.id
            ws[f"B{row_num}"] = transaction.tax_payer.name
            ws[f"C{row_num}"] = transaction.tax_payer.tin
            ws[f"D{row_num}"] = str(transaction.revenue_source)
            ws[f"E{row_num}"] = transaction.amount
            ws[f"F{row_num}"] = transaction.note
            created_naive = timezone.localtime(transaction.created).replace(tzinfo=None)
            ws[f"G{row_num}"] = created_naive

        wb.save(response)
        return response

    return render(request,
                  "transaction_list.html",
                  {"transactions": transactions,
                   "form": form})


@require_POST
def transaction_delete(request):
    if request.method == "POST":
        transaction_id = request.POST.get("transaction_id")
        transaction = get_object_or_404(RevenueTransaction, id=transaction_id)
        transaction.delete()
        messages.success(request, "Revenue Transaction successfully deleted.")
        return redirect(reverse("transaction_list"))


def revenue_categories(request):
    categories = RevenueCategory.objects.all()

    if request.method == 'POST':
        form = RevenueCategoryForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Revenue category successfully saved.")
            return redirect("revenue_categories")
        else:
            messages.error(request, "An error occured in the form below.")
    else:
        form = RevenueCategoryForm()

    return render(
        request,
        "revenue_categories.html",
        {"categories": categories,
          "form": form})


def revenue_sources(request):
    sources = RevenueSource.objects.all()

    if request.method == 'POST':
        form = RevenueSourceForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Revenue source successfully saved.")
            return redirect("revenue_sources")
        else:
            messages.error(request, "An error occured in the form below.")
    else:
        form = RevenueSourceForm()

    return render(
        request,
        "revenue_sources.html", {"sources": sources, "form": form})
